from openpyxl import load_workbook

fn = 'Microsoft Excel.xlsx'
wb = load_workbook(fn)
ws = wb['Аркуш1']
ws.append(['3', '443'])
wb.save(fn)
wb.close()
fn = 'Microsoft Excel.xlsx'
wb = load_workbook(fn)
ws = wb['Аркуш1']
ws.append(['33343', '456443'])
wb.save(fn)
wb.close()